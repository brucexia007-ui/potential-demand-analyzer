"""WBS-9: Dry Run 采样、成本估算与 Excel 解析测试"""
from __future__ import annotations

import pytest

from app.api.batch_cost import (
    calculate_sample_score,
    select_samples,
    estimate_batch_cost,
)


class TestSampleScore:
    """采样评分算法测试"""

    def test_perfect_score(self):
        """完整字段→高分"""
        score = calculate_sample_score(
            company_name="华为技术有限公司",
            demand_direction="云计算平台采购需求分析",
            skill_type="bidding",
        )
        assert score >= 0.8

    def test_short_company_name_penalty(self):
        """企业名过短→歧义惩罚"""
        score = calculate_sample_score(
            company_name="华",
            demand_direction="云计算采购",
        )
        assert score < 0.7  # 含歧义惩罚

    def test_empty_demand_direction(self):
        """空需求方向→低分"""
        score = calculate_sample_score(
            company_name="华为技术有限公司",
            demand_direction="",
        )
        assert score < 0.6

    def test_known_skill_bonus(self):
        """内置 Skill→Skill 匹配度满分"""
        score_known = calculate_sample_score(
            company_name="华为", demand_direction="云采购", skill_type="bidding"
        )
        score_unknown = calculate_sample_score(
            company_name="华为", demand_direction="云采购", skill_type="unknown"
        )
        assert score_known > score_unknown

    def test_score_range(self):
        """分数在 [0, 1] 范围内"""
        score = calculate_sample_score("测试", "测试方向")
        assert 0.0 <= score <= 1.0


class TestSelectSamples:
    """采样选择测试"""

    def test_select_top_samples(self):
        """选择最高分样本"""
        rows = [
            {"company_name": "华为技术", "demand_direction": "云计算采购"},
            {"company_name": "短", "demand_direction": "方向"},
            {"company_name": "阿里巴巴", "demand_direction": "数据中台建设"},
            {"company_name": "腾讯科技", "demand_direction": "AI平台"},
        ]
        samples = select_samples(rows, max_samples=2)
        assert len(samples) == 2
        # 高分样本排前面
        assert samples[0]["sample_score"] >= samples[1]["sample_score"]
        # "短" 应该不在样本中（得分最低）
        sample_companies = [s["company_name"] for s in samples]
        assert "短" not in sample_companies

    def test_select_more_than_rows(self):
        """样本数超过总行数→返回所有行"""
        rows = [
            {"company_name": "华为", "demand_direction": "云"},
        ]
        samples = select_samples(rows, max_samples=5)
        assert len(samples) == 1


class TestCostEstimate:
    """成本估算测试"""

    def test_estimate_from_samples(self):
        """从样本外推总成本"""
        sample_results = [
            {"tokens_used": 10000, "time_seconds": 30.0, "evidence_count": 8},
            {"tokens_used": 12000, "time_seconds": 36.0, "evidence_count": 10},
        ]
        estimate = estimate_batch_cost(sample_results, total_rows=10)
        assert estimate["total_rows"] == 10
        assert estimate["sample_count"] == 2
        # 外推: avg=11000 tokens/sample, ×10 = 110000
        assert estimate["estimated_total_tokens"] == 110000
        assert estimate["estimated_total_time_minutes"] > 0
        assert estimate["confidence"] == "medium"
        assert estimate["monetary_cost"] == {
            "status": "UNAVAILABLE",
            "amount": None,
            "currency": None,
            "reason": "当前模型与搜索供应商没有统一价目表，禁止伪造金额估算。",
        }
        assert "estimated_cost_rmb" not in estimate
        assert estimate["estimate_basis"] == "Skill 声明预算的确定性规划外推，Dry Run 不调用外部 Provider。"

    def test_empty_samples(self):
        """无样本→零估算"""
        estimate = estimate_batch_cost([], total_rows=10)
        assert estimate["estimated_total_tokens"] == 0
        assert estimate["confidence"] == "low"

    def test_single_sample_low_confidence(self):
        """单样本→低置信度"""
        sample_results = [
            {"tokens_used": 5000, "time_seconds": 15.0, "evidence_count": 5},
        ]
        estimate = estimate_batch_cost(sample_results, total_rows=20)
        assert estimate["confidence"] == "low"
        assert estimate["estimated_total_tokens"] == 100000  # 5000 × 20


class TestExcelParser:
    """Excel 解析测试"""

    def test_parse_basic_xlsx(self):
        """解析标准 .xlsx 文件"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl 未安装")

        from app.api.batch_parser import parse_excel_to_rows

        # 创建测试 Excel
        import io
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["公司名称", "需求方向"])
        ws.append(["华为", "云计算"])
        ws.append(["阿里", "AI平台"])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        result = parse_excel_to_rows(buf.read(), "test.xlsx")
        assert result["source_row_count"] == 2
        assert result["candidate_rows"][0]["company_name"] == "华为"
        assert result["candidate_rows"][1]["demand_direction"] == "AI平台"

    def test_parse_empty_excel(self):
        """空 Excel→抛异常"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl 未安装")

        from app.api.batch_parser import parse_excel_to_rows, CsvParseError

        import io
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        with pytest.raises(CsvParseError, match="至少需要包含表头"):
            parse_excel_to_rows(buf.read(), "empty.xlsx")

    def test_parse_excel_english_columns(self):
        """英文列名 Excel"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl 未安装")

        from app.api.batch_parser import parse_excel_to_rows

        import io
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["company_name", "demand_direction"])
        ws.append(["Apple", "iPhone supply chain"])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        result = parse_excel_to_rows(buf.read(), "eng.xlsx")
        assert result["source_row_count"] == 1
        assert result["candidate_rows"][0]["company_name"] == "Apple"

    def test_parse_excel_file_too_large(self):
        """超大 Excel→抛异常"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl 未安装")

        from app.api.batch_parser import parse_excel_to_rows, CsvParseError

        # 创建一个 >10MB 的假文件
        large_content = b"x" * (11 * 1024 * 1024)
        with pytest.raises(CsvParseError, match="文件过大"):
            parse_excel_to_rows(large_content, "large.xlsx")
