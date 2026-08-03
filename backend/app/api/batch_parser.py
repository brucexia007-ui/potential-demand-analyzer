"""CSV 解析：将上传的 CSV 文件解析为 BatchTaskInput 列表

支持中英文列名智能匹配，跳过空行，行数上限 1000。
"""

import csv
import io
from typing import Optional

from app.api.batch_template_service import BatchTemplateService

# 列名别名映射 → 标准列名
COLUMN_ALIASES: dict[str, str] = {
    # company_name 别名
    "company_name": "company_name",
    "公司名称": "company_name",
    "企业名称": "company_name",
    "企业": "company_name",
    "公司": "company_name",
    "客户名称": "company_name",
    "客户": "company_name",
    "company": "company_name",
    "name": "company_name",
    # demand_direction 别名
    "demand_direction": "demand_direction",
    "需求方向": "demand_direction",
    "查询方向": "demand_direction",
    "需求": "demand_direction",
    "分析方向": "demand_direction",
    "调研方向": "demand_direction",
    "direction": "demand_direction",
    "demand": "demand_direction",
    # v3.1: 行业别名
    "industry": "industry",
    "行业": "industry",
    "所属行业": "industry",
    "行业类型": "industry",
    "行业分类": "industry",
    # v3.1: 地区别名
    "region": "region",
    "地区": "region",
    "所在地区": "region",
    "省份": "region",
    "城市": "region",
    "区域": "region",
    "地点": "region",
    # 主体消歧与能力档案（全部非必填）
    "official_website": "official_website",
    "官网": "official_website",
    "企业官网": "official_website",
    "unified_social_credit_code": "unified_social_credit_code",
    "统一社会信用代码": "unified_social_credit_code",
    "信用代码": "unified_social_credit_code",
    "capability_profile_id": "capability_profile_id",
    "企业能力档案ID": "capability_profile_id",
}

# v3.1: 可选列（不是必须的）
OPTIONAL_STANDARD_COLUMNS = {"industry", "region"}

MAX_ROWS = 1000
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5 MB


class CsvParseError(Exception):
    """CSV 解析错误"""
    pass


def _normalize_headers(
    headers: list[str],
    *,
    required_columns: tuple[str, ...] = ("company_name", "demand_direction"),
) -> dict[str, str]:
    """将 CSV/Excel 表头映射到标准列名

    Returns:
        {"company_name": "原始列名", "demand_direction": "原始列名", ...}
        可选列（industry/region）仅在存在时返回。

    Raises:
        CsvParseError: 缺少必填列（company_name, demand_direction）
    """
    mapping: dict[str, str] = {}
    for idx, h in enumerate(headers):
        # 去掉 BOM 和空白
        h_clean = h.strip().lstrip("﻿")
        standard = COLUMN_ALIASES.get(h_clean)
        if standard and standard not in mapping:
            mapping[standard] = h_clean

    if "company_name" in required_columns and "company_name" not in mapping:
        raise CsvParseError(
            "CSV 缺少企业名称列。请确保包含以下列名之一：company_name、公司名称、企业名称、企业、公司"
        )
    if "demand_direction" in required_columns and "demand_direction" not in mapping:
        raise CsvParseError(
            "CSV 缺少需求方向列。请确保包含以下列名之一：demand_direction、需求方向、查询方向、需求"
        )

    return mapping


def _template_contract(template_id: str, version: int | None) -> tuple[str, int, tuple[str, ...]]:
    template = BatchTemplateService().get(template_id)
    if version is not None and version != template.version:
        raise CsvParseError(
            f"模板版本不受支持：{template_id}@{version}，当前版本为 {template.version}"
        )
    required = tuple(field.key for field in template.fields if field.required)
    return template.template_id, template.version, required


def _parse_tabular_rows(
    all_rows: list[list[str]],
    *,
    template_id: str,
    template_version: int,
    required_columns: tuple[str, ...],
) -> dict:
    if len(all_rows) < 2:
        raise CsvParseError("文件至少需要包含表头行和一行数据")
    raw_headers = all_rows[0]
    column_map = _normalize_headers(raw_headers, required_columns=required_columns)
    header_to_idx = {header.strip().lstrip("﻿"): index for index, header in enumerate(raw_headers)}

    def value(row: list[str], key: str) -> str:
        original = column_map.get(key)
        if original is None:
            return ""
        index = header_to_idx[original]
        return (row[index] if index < len(row) else "").strip()

    candidate_rows: list[dict] = []
    for source_row_index, row in enumerate(all_rows[1:], start=2):
        if not any(cell.strip() for cell in row):
            continue
        company = value(row, "company_name")
        if company.startswith("【示例】"):
            continue
        demand = value(row, "demand_direction")
        if template_id == "opportunity_discovery" and not demand:
            demand = "自动发现潜在需求与商机线索"
        candidate: dict = {
            "source_row_index": source_row_index,
            "company_name": company or None,
            "demand_direction": demand or None,
        }
        for key in ("industry", "region", "capability_profile_id"):
            item = value(row, key)
            if item:
                candidate[key] = item
        disambiguation = {
            key: item
            for key in ("official_website", "unified_social_credit_code")
            if (item := value(row, key))
        }
        if disambiguation:
            candidate["disambiguation"] = disambiguation
        candidate_rows.append(candidate)
        if len(candidate_rows) >= MAX_ROWS:
            break
    if not candidate_rows:
        raise CsvParseError("文件中未找到非空数据行")
    return {
        "template_id": template_id,
        "template_version": template_version,
        "headers": [header.strip() for header in raw_headers],
        "source_row_count": len(candidate_rows),
        "candidate_rows": candidate_rows,
        "preview_candidates": candidate_rows[:5],
    }


def parse_csv_to_rows(file_content: bytes, filename: str = "") -> dict:
    """解析 CSV 文件内容为任务行列表

    Args:
        file_content: CSV 文件原始字节
        filename: 文件名（仅用于错误提示）

    Returns:
        {
            "filename": str,
            "row_count": int,
            "headers": [str, ...],
            "rows": [{ "company_name": str, "demand_direction": str }, ...],
            "preview": [...],  # 前 5 行
        }

    Raises:
        CsvParseError: 解析失败
    """
    if len(file_content) > MAX_FILE_SIZE:
        raise CsvParseError(f"文件过大（最大 {MAX_FILE_SIZE // 1024 // 1024} MB）")

    # 尝试 UTF-8 解码，失败则尝试 GBK
    text: Optional[str] = None
    for encoding in ("utf-8", "utf-8-sig", "gbk", "gb2312"):
        try:
            text = file_content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue

    if text is None:
        raise CsvParseError("无法识别文件编码，请使用 UTF-8 或 GBK 编码")

    # 用 csv.Sniffer 检测分隔符
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # 默认逗号

    reader = csv.reader(io.StringIO(text), dialect)
    rows_raw = list(reader)

    if len(rows_raw) < 2:
        raise CsvParseError("CSV 文件至少需要包含表头行和一行数据")

    template_id = "standard_research"
    declared_version: int | None = None
    if rows_raw and rows_raw[0] and rows_raw[0][0].strip().lstrip("﻿") == "__kanyikan_template__":
        if len(rows_raw[0]) < 3 or len(rows_raw) < 3:
            raise CsvParseError("CSV 模板元数据不完整")
        template_id = rows_raw[0][1].strip()
        try:
            declared_version = int(rows_raw[0][2].strip())
        except ValueError as error:
            raise CsvParseError("CSV 模板版本必须为整数") from error
        rows_raw = rows_raw[1:]
    try:
        template_id, template_version, required_columns = _template_contract(template_id, declared_version)
    except LookupError as error:
        raise CsvParseError(str(error)) from error

    raw_headers = rows_raw[0]
    try:
        _normalize_headers(raw_headers, required_columns=required_columns)
    except CsvParseError as e:
        # 如果 Sniffer 检测了错误的分隔符，尝试用逗号重新解析
        if dialect.delimiter != ",":
            reader2 = csv.reader(io.StringIO(text))
            rows_raw2 = list(reader2)
            if len(rows_raw2) >= 2:
                try:
                    _normalize_headers(rows_raw2[0], required_columns=required_columns)
                    rows_raw = rows_raw2
                    dialect = csv.excel
                except CsvParseError:
                    raise e
            else:
                raise e
        else:
            raise e

    parsed = _parse_tabular_rows(
        rows_raw,
        template_id=template_id,
        template_version=template_version,
        required_columns=required_columns,
    )
    return {
        "filename": filename or "upload.csv",
        **parsed,
    }


# ── WBS-9.8 Excel 解析 ─────────────────────────────────────────────────


MAX_EXCEL_FILE_SIZE = 10 * 1024 * 1024  # 10 MB


def parse_excel_to_rows(file_content: bytes, filename: str = "") -> dict:
    """解析 Excel (.xlsx/.xls) 文件内容为任务行列表

    只读取第一个工作表，第一行为表头。
    列名映射规则与 CSV 解析一致。

    Args:
        file_content: Excel 文件原始字节
        filename: 文件名（仅用于错误提示）

    Returns:
        与 parse_csv_to_rows 相同格式

    Raises:
        CsvParseError: 解析失败
        ImportError: openpyxl 未安装
    """
    if len(file_content) > MAX_EXCEL_FILE_SIZE:
        raise CsvParseError(f"文件过大（最大 {MAX_EXCEL_FILE_SIZE // 1024 // 1024} MB）")

    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl 未安装，请运行: pip install openpyxl")

    from io import BytesIO

    try:
        wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True, data_only=True)
    except Exception as e:
        raise CsvParseError(f"无法打开 Excel 文件: {e}")

    try:
        ws = wb["导入数据"] if "导入数据" in wb.sheetnames else wb.active
        if ws is None:
            raise CsvParseError("Excel 文件中没有工作表")

        template_id = "standard_research"
        declared_version: int | None = None
        if "_template_meta" in wb.sheetnames:
            metadata_sheet = wb["_template_meta"]
            metadata = {
                str(row[0]).strip(): row[1]
                for row in metadata_sheet.iter_rows(values_only=True)
                if row and len(row) >= 2 and row[0] is not None
            }
            template_id = str(metadata.get("template_id") or "").strip()
            try:
                declared_version = int(metadata.get("template_version"))
            except (TypeError, ValueError) as error:
                raise CsvParseError("Excel 模板版本必须为整数") from error
        try:
            template_id, template_version, required_columns = _template_contract(template_id, declared_version)
        except LookupError as error:
            raise CsvParseError(str(error)) from error

        # 读取所有行（只读模式下 ws.rows 是生成器）
        all_rows: list[list[str]] = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            str_row = [str(cell).strip() if cell is not None else "" for cell in row]
            # 跳过完全空行
            if any(str_row):
                all_rows.append(str_row)

        wb.close()

        parsed = _parse_tabular_rows(
            all_rows,
            template_id=template_id,
            template_version=template_version,
            required_columns=required_columns,
        )
        return {
            "filename": filename or "upload.xlsx",
            **parsed,
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass
