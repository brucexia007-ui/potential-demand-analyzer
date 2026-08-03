"""批量导入标准模板目录与 XLSX/CSV 生成服务。"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


TemplateFormat = Literal["xlsx", "csv"]


@dataclass(frozen=True)
class BatchTemplateField:
    key: str
    label: str
    required: bool
    description: str
    example: str


@dataclass(frozen=True)
class BatchTemplateDefinition:
    template_id: str
    version: int
    name: str
    description: str
    fields: tuple[BatchTemplateField, ...]


@dataclass(frozen=True)
class GeneratedBatchTemplate:
    filename: str
    media_type: str
    content: bytes


_COMMON_OPTIONAL_FIELDS = (
    BatchTemplateField("official_website", "官网", False, "用于企业主体消歧；非必填", "https://example.com"),
    BatchTemplateField("unified_social_credit_code", "统一社会信用代码", False, "用于企业主体精确消歧；非必填", "91310000XXXXXXXXXX"),
    BatchTemplateField("region", "地区", False, "省/市/国家或主要经营区域；非必填", "上海"),
    BatchTemplateField("industry", "行业", False, "目标企业所属行业；非必填", "制造业"),
)

_TEMPLATES: dict[str, BatchTemplateDefinition] = {
    "standard_research": BatchTemplateDefinition(
        template_id="standard_research",
        version=1,
        name="标准客户研究模板",
        description="适合已明确研究方向的批量客户研究。",
        fields=(
            BatchTemplateField("company_name", "企业名称", True, "目标客户企业名称", "某科技有限公司"),
            BatchTemplateField("demand_direction", "需求方向", True, "希望重点研究的问题或业务方向", "数据治理与智能客服"),
            *_COMMON_OPTIONAL_FIELDS,
        ),
    ),
    "opportunity_discovery": BatchTemplateDefinition(
        template_id="opportunity_discovery",
        version=1,
        name="自动线索发现模板",
        description="只需企业名称；系统将结合选定企业能力档案自动研究潜在线索。",
        fields=(
            BatchTemplateField("company_name", "企业名称", True, "唯一必填项：目标客户企业名称", "某制造集团有限公司"),
            *_COMMON_OPTIONAL_FIELDS,
            BatchTemplateField("capability_profile_id", "企业能力档案ID", False, "指定我方能力档案；留空时使用创建批次时选择的档案", ""),
        ),
    ),
}


class BatchTemplateService:
    def list_templates(self) -> tuple[BatchTemplateDefinition, ...]:
        return tuple(_TEMPLATES.values())

    def get(self, template_id: str) -> BatchTemplateDefinition:
        template = _TEMPLATES.get(template_id)
        if template is None:
            raise LookupError("批量导入模板不存在")
        return template

    def generate(self, *, template_id: str, file_format: TemplateFormat) -> GeneratedBatchTemplate:
        template = self.get(template_id)
        if file_format == "xlsx":
            content = self._xlsx(template)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif file_format == "csv":
            content = self._csv(template)
            media_type = "text/csv; charset=utf-8"
        else:
            raise ValueError("模板格式仅支持 xlsx 或 csv")
        return GeneratedBatchTemplate(
            filename=f"kanyikan_{template.template_id}_v{template.version}.{file_format}",
            media_type=media_type,
            content=content,
        )

    @staticmethod
    def _xlsx(template: BatchTemplateDefinition) -> bytes:
        workbook = Workbook()
        data_sheet = workbook.active
        data_sheet.title = "导入数据"
        labels = [field.label for field in template.fields]
        data_sheet.append(labels)
        data_sheet.freeze_panes = "A2"
        data_sheet.auto_filter.ref = f"A1:{get_column_letter(len(labels))}1"
        required_fill = PatternFill("solid", fgColor="FFF1C2")
        optional_fill = PatternFill("solid", fgColor="E8F1FF")
        for index, field in enumerate(template.fields, start=1):
            cell = data_sheet.cell(row=1, column=index)
            cell.font = Font(bold=True, color="111111")
            cell.fill = required_fill if field.required else optional_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            data_sheet.column_dimensions[get_column_letter(index)].width = min(42, max(14, len(field.label) * 2 + 6))

        guide = workbook.create_sheet("填写说明")
        guide.append([template.name, template.description])
        guide.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        guide["A1"].font = Font(bold=True, size=14)
        guide.append(["字段键", "列名", "是否必填", "填写说明", "示例"])
        for field in template.fields:
            guide.append([field.key, field.label, "必填" if field.required else "非必填", field.description, field.example])
        guide.freeze_panes = "A3"
        for index, width in enumerate((30, 24, 12, 58, 36), start=1):
            guide.column_dimensions[get_column_letter(index)].width = width
        for cell in guide[2]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E6E6E6")

        example = workbook.create_sheet("填写示例")
        example.append(labels)
        example.append([field.example for field in template.fields])
        for index in range(1, len(labels) + 1):
            example.column_dimensions[get_column_letter(index)].width = data_sheet.column_dimensions[get_column_letter(index)].width

        metadata = workbook.create_sheet("_template_meta")
        metadata.sheet_state = "hidden"
        metadata.append(["template_id", template.template_id])
        metadata.append(["template_version", template.version])
        metadata.append(["mode", template.template_id])
        metadata.append(["header_sheet", "导入数据"])

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def _csv(template: BatchTemplateDefinition) -> bytes:
        output = io.StringIO(newline="")
        writer = csv.writer(output)
        writer.writerow(["__kanyikan_template__", template.template_id, template.version])
        writer.writerow([field.label for field in template.fields])
        writer.writerow([f"【示例】{field.example}" if index == 0 else field.example for index, field in enumerate(template.fields)])
        return ("\ufeff" + output.getvalue()).encode("utf-8")
