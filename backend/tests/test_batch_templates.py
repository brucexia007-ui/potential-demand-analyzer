"""标准与自动线索发现批量模板必须自描述、可下载且只强制必要字段。"""
from __future__ import annotations

import csv
import io

from openpyxl import load_workbook


def test_template_catalog_keeps_disambiguation_fields_optional() -> None:
    from app.api.batch_template_service import BatchTemplateService

    templates = {item.template_id: item for item in BatchTemplateService().list_templates()}

    assert set(templates) == {"standard_research", "opportunity_discovery"}
    assert [field.key for field in templates["standard_research"].fields if field.required] == [
        "company_name", "demand_direction",
    ]
    assert [field.key for field in templates["opportunity_discovery"].fields if field.required] == ["company_name"]
    assert all(
        not field.required
        for field in templates["opportunity_discovery"].fields
        if field.key in {"official_website", "unified_social_credit_code", "region", "industry"}
    )


def test_xlsx_template_contains_data_guide_example_and_hidden_version_metadata() -> None:
    from app.api.batch_template_service import BatchTemplateService

    generated = BatchTemplateService().generate(template_id="opportunity_discovery", file_format="xlsx")
    workbook = load_workbook(io.BytesIO(generated.content), data_only=True)

    assert generated.filename == "kanyikan_opportunity_discovery_v1.xlsx"
    assert workbook.sheetnames == ["导入数据", "填写说明", "填写示例", "_template_meta"]
    assert workbook["_template_meta"].sheet_state == "hidden"
    assert workbook["_template_meta"]["B1"].value == "opportunity_discovery"
    assert workbook["_template_meta"]["B2"].value == 1
    assert workbook["导入数据"]["A1"].value == "企业名称"
    assert workbook["导入数据"].max_row == 1
    assert workbook["填写示例"].max_row == 2


def test_csv_template_has_machine_readable_version_and_separate_example_marker() -> None:
    from app.api.batch_template_service import BatchTemplateService

    generated = BatchTemplateService().generate(template_id="standard_research", file_format="csv")
    rows = list(csv.reader(io.StringIO(generated.content.decode("utf-8-sig"))))

    assert rows[0] == ["__kanyikan_template__", "standard_research", "1"]
    assert rows[1][:2] == ["企业名称", "需求方向"]
    assert rows[2][0].startswith("【示例】")
